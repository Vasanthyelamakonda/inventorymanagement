from django.shortcuts import get_object_or_404, render,redirect
from django.http import HttpResponse
import mysql.connector
from .models import Warehouse, Dailysheet, Brands, Dailysales, Bills, Brand, Employee, Accountant
from django.core.files.storage import FileSystemStorage
from pathlib import Path
from wsgiref.util import FileWrapper #django >1.8
from mimetypes import MimeTypes
from datetime import datetime
from random import randint
import smtplib,ssl
from random import randint
import openpyxl
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.urls import reverse 

import pandas as pd

# Create your views here.
def index(request):
    return render(request,'index.html')
def regemp(request):
    if request.method == 'POST':        
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
        ) 
        mycursor =conn.cursor()
        #retrive post details       
        username=request.POST['username']
        password=request.POST['password']     
        email=request.POST['email']
        mobile=request.POST['mobile']
        dob=request.POST['dob']
        mycursor.execute("insert into employee (username,password,email,mobile,dob) values('"+username+"','"+password+"','"+email+"','"+mobile+"','"+dob+"')")
        conn.commit()
        return redirect('regemp')
    elif "username" in request.session:
        
        username=request.session['username']
    
        return render(request,'regemp.html')
    else:
         return render(request,'adminlogin.html')
     
def regacc(request):
    if request.method == 'POST':        
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
        ) 
        mycursor =conn.cursor()
        #retrive post details       
        username=request.POST['username']
        password=request.POST['password']     
        email=request.POST['email']
        mobile=request.POST['mobile']
        dob=request.POST['dob']
          
        mycursor.execute("insert into accountant (username,password,email,mobile,dob) values('"+username+"','"+password+"','"+email+"','"+mobile+"','"+dob+"')")
        conn.commit()
        return redirect('regacc')
    elif "username" in request.session:
        
        username=request.session['username']
    
        return render(request,'regacc.html')
    else:
         return render(request,'adminlogin.html')
        
def adminlogin(request):

    if request.method == 'POST':
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
        
        )
        mycursor = conn.cursor()
        #retrive post details       
        
        username=request.POST['username']
        password=request.POST['password']     
           
        mycursor.execute("select * from admin where username='"+username+"' and password='"+password+"'")
        result=mycursor.fetchone()
        if(result!=None):
            request.session["username"]=username
            return render(request, 'admindashboard.html')
        else:
            return render(request,'adminlogin.html',{'status':'invalid credentials'})  
    else:
        return render(request,'adminlogin.html')
    
    
def adminlogout(request):
    try:
        del request.session['username']
        request.session.modified= True
        return render(request,'adminlogin.html')
    except KeyError:
        return redirect('adminlogin')
    
    
def emplogout(request):
    try:
        del request.session['username']
        request.session.modified= True
        return render(request,'emplogin.html')
    except KeyError:
        return redirect('emplogin')
    
def acclogout(request):
    try:
        del request.session['username']
        request.session.modified= True
        return render(request,'acclogin.html')
    except KeyError:
        return redirect('acclogin')


  
def emplogin(request):

    if request.method == 'POST':
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
        
        )
        mycursor = conn.cursor()
        #retrive post details       
        
        username=request.POST['username']
        password=request.POST['password']     
           
        mycursor.execute("select * from employee where username='"+username+"' and password='"+password+"'")
        result=mycursor.fetchone()
        if(result!=None):
            request.session["username"]=username
            return render(request, 'empdashboard.html')
            
        else:
            return render(request,'emplogin.html')  
    else:
        return render(request,'emplogin.html') 

def acclogin(request):

    if request.method == 'POST':
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
        
        )
        mycursor = conn.cursor()
        #retrive post details       
        
        username=request.POST['username']
        password=request.POST['password']     
           
        mycursor.execute("select * from accountant where username='"+username+"' and password='"+password+"'")
        result=mycursor.fetchone()
        if(result!=None):
            return render(request, 'accdashboard.html', {"username" : username})
        else:
            return render(request,'acclogin.html',{'status':'invalid credentials'})  
    else:
        return render(request,'acclogin.html')   

def about(request):
    return render(request,"about.html")

def pricing(request):
    return render(request,"pricing.html")

def employee(request):
    return render(request,"employee.html")

def contact(request):
    return render(request,"contact.html")

def admindashboard(request):
    if "username" in request.session:
        username=request.session['username']
        return render(request, "admindashboard.html")
    else:
        return render(request,'adminlogin.html')

def empdashboard(request):
    return render(request, "empdashboard.html")

def accdashboard(request):
    return render(request, "accdashboard.html")

def dailysheetacc(request):
    if request.method=='POST':
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        query = conn.cursor()
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select d_id,date,brand_name,ML,Opening_bal,plant,ob_total,Closing_bal,sales,MRP,amount from daily_sheet where date between "'+fromdate+'" and "'+todate+'"  ')
        searchresult=query.fetchall()
        d=[]
        a=[]
        R_sale=0 
        Exp=0
        Bal=0
        cashob=0
        total=0
        recash=0
        cash=0
      
        for row in searchresult: 
            obj= Dailysheet()
            obj.date=row[1] 
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
            R_sale=R_sale+int(obj.amount)    
            d.append(obj)
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select  date, retail_sale,expenditure,balance,cashob,total,recash,handcash from dailysales where date between "'+fromdate+'" and "'+todate+'" ')
        result=query.fetchall()

        for row in result:
            obj=Dailysales()
            obj.date=row[0]
            obj.retail_sale=row[1]
            obj.expenditure=row[2]
            obj.balance=row[3]
            obj.cashob=row[4]
            obj.total=row[5]
            obj.recash=row[6]
            obj.handcash=row[7]
            Exp=Exp+int(obj.expenditure)
            Bal=Bal+int(obj.balance)
            cashob=cashob+int(obj.cashob)
            total=total+int(obj.total)
            recash=recash+int(obj.recash)
            cash=cash+int(obj.handcash)

            a.append(obj)
            

        return render(request,'dailysheetacc.html',{"data":d,"Rsale" : R_sale, "EXP":Exp,"bal":Bal,"ob":cashob,"tt":total,"cash":recash,"csh":cash})
        
    elif  "username" in request.session:
        
        username=request.session['username']
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
            
        mycursor = conn.cursor()
        mycursor.execute("select * from daily_sheet")
        result = mycursor.fetchall()
        d=[] 
        R_sale=0
        for row in result:
            obj= Dailysheet()
            obj.d_id=row[0]
            obj.date=row[1]
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
           
            R_sale=R_sale+int(obj.amount)  
          
            d.append(obj)
        #print(e)
        
        return render(request, 'dailysheetacc.html', {'data': d, 'Rsale' : R_sale})

    else:
        return render(request,'acclogin.html')
    



def brandsandpriceadmin(request):
    if 'q' in request.GET:
        q=request.GET['q']
        brands=Brand.objects.filter(brand_name__icontains=q)
    else:
        brands=Brand.objects.all()
    context={
        'brands':brands
    }
    return render(request,'brandsandpriceadmin.html',context)
        
    
    
def brandsandpriceemp(request):
    if 'q' in request.GET:
        q=request.GET['q']
        brands=Brand.objects.filter(brand_name__icontains=q)
    else:
        brands=Brand.objects.all()
    context={
        'brands':brands
    }
    return render(request,'brandsandpriceemp.html',context)
    
            
            

def challanandbills(request):
    
    
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
    )
    mycursor=conn.cursor()
    mycursor.execute("select * from challan_details")
    result=mycursor.fetchall()
    files=[]
    for x in result:
        f=Bills()
        f.cid=x[0]
        f.date=x[1]
        f.bill=x[2]   
        files.append(f) 
    return render(request,'challanandbills.html',{"files":files})


def challanandbillsacc(request):
    
    
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
    )
    mycursor=conn.cursor()
    mycursor.execute("select * from challan_details")
    result=mycursor.fetchall()
    files=[]
    for x in result:
        f=Bills()
        f.cid=x[0]
        f.date=x[1]
        f.bill=x[2]   
        files.append(f) 
    return render(request,'challanandbillsacc.html',{"files":files})
    
    
def load(request):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
    )
    mycursor = mydb.cursor()
    #retrive post details       
    mycursor.execute("select * from  challan_details")
    result=mycursor.fetchall()
    files=[]
    for x in result:
        f=Bills()
        f.name=x[0]      
        files.append(f)    
    return render(request,'load.html',{"files":files})
   

 
def others(request): 
    if "username" in request.session:
        username=request.session['username']
    
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from employee")
        result = mycursor.fetchall()
        e=[] 
        for row in result:
            obj=Employee()
            obj.emp_id=row[0]
            obj.username=row[1]
            obj.email=row[3]
            obj.dob=row[4] 
            obj.mobile = row[5]
            e.append(obj)   
        mycursor.execute("select * from accountant")
        res = mycursor.fetchall()
        a=[] 
        for row in res:
            obj=Accountant()
            obj.acc_id=row[0]
            obj.username=row[1]
            obj.email=row[3]
            obj.dob=row[4] 
            obj.mobile = row[5]
            a.append(obj)   
        
        return render(request, 'others.html', {'accountant': a,'employee': e})
    else:
            return render(request,'adminlogin.html')  





def members(request):
    return render(request, "members.html")

def addsheet(request):
    if 'button' in request.POST:
        getdate=request.POST['date']
        bname=request.POST['brandname']
        ml=request.POST['ml']
        what=request.POST['form_id']

        if what == 'getdata':
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="",
                database="inventorymanagement"
            )
            mycursor = mydb.cursor()
            #retrive post details       
            mycursor.execute("select * from daily_sheet where date='"+getdate+"' and brand_name='"+bname+"' and ML='"+ml+"' ")
            result=mycursor.fetchone()
            obj= Dailysheet()
            if(result!=None):              
                obj.date=result[1] 
                obj.brand_name	 = result[2]
                obj.ML = result[3]
                obj.Opening_bal	= result[4]
                obj.plant= result[5]
                obj.ob_total= result[6]
                obj.Closing_bal	= result[7]
                obj.sales= result[8]
                obj.MRP= result[9]
                obj.amount= result[10]
            print(obj)
            result=Brand.objects.values('brand_name').distinct()
            return render(request, 'addsheet.html',{'previous':obj,"Brand":result})
   
    
        else:
            conn = mysql.connector.connect(
                host = 'localhost',
                user = 'root',
                password = '',
                database = 'inventorymanagement'
            )
            query = conn.cursor()
            date=request.POST['date']
            bname=request.POST['brandname']
            ml=request.POST['ml']
            ob=request.POST['Openingbal']
            Plant=request.POST['plant']
            total=request.POST['total']
            cb=request.POST['Closingbal']
            sales=request.POST['sale']
            mrp=request.POST['rate']
           
            query.execute("insert into daily_sheet (date,brand_name,ML,Opening_bal, plant, ob_total, Closing_bal, sales, MRP)  values('"+date+"','"+bname+"','"+ml+"','"+ob+"','"+Plant+"','"+total+"','"+cb+"','"+sales+"','"+mrp+"') ")
            conn.commit()
            return redirect('dailysheetemp')
    
    elif 'Add' in request.POST:

        conn = mysql.connector.connect(
                host = 'localhost',
                user = 'root',
                password = '',
                database = 'inventorymanagement'
            

        )
        query = conn.cursor()
        date=request.POST['date']
        bname=request.POST['brandname']
        ml=request.POST['ml']
        ob=request.POST['Openingbal']
        Plant=request.POST['plant']
        total=request.POST['total']
        cb=request.POST['Closingbal']
        sales=request.POST['sale']
        mrp=request.POST['rate']
        
        query.execute("insert into daily_sheet (date,brand_name,ML,Opening_bal, plant, ob_total, Closing_bal, sales, MRP)  values('"+date+"','"+bname+"','"+ml+"','"+ob+"','"+Plant+"','"+total+"','"+cb+"','"+sales+"','"+mrp+"') ")
        conn.commit()
        return redirect('dailysheetemp')
    else:
        
        result=Brand.objects.values('brand_name').distinct()
        
        return render(request,'addsheet.html',{"Brand":result}) 
       
        

def addbrands(request):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        bn=request.POST['brandnumber']
        bname=request.POST['brandname']
        ml=request.POST['ML']
        price=request.POST['issue_price']
        mrp=request.POST['MRP']

        if Brand.objects.filter(brand_no=bn).exclude(brand_name=bname).exists():
            status = 'Brand already exists with a different name!'
            return render(request, 'addbrands.html', {'status': status})
        # check if the brand already exists with the same brand number and brand name
        elif Brand.objects.filter(brand_no=bn, brand_name=bname, ML=ml).exists():
            status = 'Brand already exists!'
            return render(request, 'addbrands.html', {'status': status})
        else:
            brand = Brand(brand_no=bn, brand_name=bname, ML=ml, issue_price=price, MRP=mrp)
            brand.save()
            success_msg = 'Brand added successfully!'
            return redirect('brandsandpriceadmin')

    
        #query.execute("insert into adminapp_brand (brand_no,brand_name,ML,issue_price,MRP)  values('"+bn+"', '"+bname+"','"+ml+"','"+price+"','"+mrp+"')")
        #conn.commit()
        #return redirect('brandsandpriceadmin')
        
    else:
        conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventorymanagement"
            
        )
        
        mycursor = conn.cursor()
        mycursor.execute("select * from adminapp_brand")
        result=mycursor.fetchall()
        b=[]
        for row in result:
            obj=Brands()        
            obj.bid=row[0]
            obj.brand_no=row[1]
            obj.brand_name=row[2]
            obj.ML=row[3]
            obj.issue_price=row[4]
            obj.MRP=row[5]
            b.append(obj)        
        return render(request, 'addbrands.html',{'brands':b})
        
        
     

def stockmanagementemp(request):
    if request.method=='POST':
        conn= mysql.connector.connect(
        host='localhost',
        user='root',
            password='',
            database='inventorymanagement'
    )
    
    elif "username" in request.session:
        
        username=request.session['username'] 
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select w.w_id,w.date,w.brand_no,w.brand_name,w.size,w.quantity_boxes,ab.issue_price,w.quantity_boxes*ab.issue_price as amount from warehouse w inner join adminapp_brand ab on w.brand_name=ab.brand_name")
        result = mycursor.fetchall()
        w=[] 
        for row in result:
            obj=Warehouse()
            obj.w_id=row[0]
            obj.date=row[1] 
            obj.brand_no=row[2] 
            obj.brand_name = row[3]
            obj.size = row[4]
            obj.quantity_boxes= row[5]
            obj.issue_price=row[6]
            obj.amount=row[7]
        
            w.append(obj)
        #print(e)
        
        return render(request, 'stockmanagementemp.html', {'warehouse': w})
    else:
         return render(request, 'emplogin.html')
        


def stockmanagementadmin(request):
    
    
    if "username" in request.session:
        username=request.session['username']
    
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select w.w_id,w.date,w.brand_no,w.brand_name,w.size,w.quantity_boxes,ab.issue_price,w.quantity_boxes*ab.issue_price as amount from warehouse w inner join adminapp_brand ab on w.brand_name=ab.brand_name")
        result = mycursor.fetchall()
        w=[] 
        for row in result:
            obj=Warehouse()
            obj.w_id=row[0]
            obj.date=row[1]
            obj.brand_no=row[2] 
            obj.brand_name = row[3]
            obj.size = row[4]
            obj.quantity_boxes= row[5]
            obj.issue_price=row[6]
            obj.amount=row[7]

            w.append(obj)   
        #print(e)
        
        return render(request, 'stockmanagementadmin.html', {'warehouse': w})
    else:
        return render(request,'adminlogin.html')

def addstock(request):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password ='',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        sdate=request.POST['date'] 
        brandname=request.POST['brandname']
        size=request.POST['size']   
        boxes=request.POST['quantityincases']  
        query.execute("select brand_no from adminapp_brand where brand_name='"+brandname+"'")
        res=query.fetchone()
        print(res[0])
        query.execute("insert into warehouse (date,brand_no,brand_name,size,quantity_boxes)  values('"+sdate+"','"+str(res[0])+"','"+brandname+"','"+size+"','"+boxes+"') ")
        conn.commit()
        return redirect('stockmanagementemp')
       
    else:
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        query.execute("select * from warehouse")
        result= query.fetchall()
        brands=[]
        for row in result:
            obj=Warehouse()
            obj.w_id=row[0]
            obj.date=row[1]
            obj.brand_no=row[2] 
            obj.brand_name = row[3]
            obj.size = row[4]
            obj.quantity_boxes= row[5]      
            brands.append(obj)
            result=Brand.objects.values('brand_name').distinct()
            res=Brand.objects.values('brand_no').distinct()
        return render(request,'addstock.html',{"brands":brands,"Br":result,"Brand":res})
        

   
def editstock(request,w_id):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
       
        brandnumber = request.POST['brandnumber']
        brandname = request.POST['brandname']
        brandsize = request.POST['size']
        quantity_boxes = request.POST['quantityincases']
       
        
        
        query.execute("update warehouse set brand_no ='"+brandnumber+"',brand_name='"+brandname+"',size='"+brandsize+"',quantity_boxes='"+quantity_boxes+"' where w_id='"+w_id+"'")
        conn.commit()
        return redirect(stockmanagementemp)
    elif "username" in request.session:
        
        username=request.session['username']
    
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from warehouse where w_id='"+w_id+"'")
        result = mycursor.fetchall()
        w=[] 
        for row in result:
            obj= Warehouse()
            obj.w_id=row[0]
            obj.date=row[1]
            obj.brand_no=row[2] 
            obj.brand_name= row[3]
            obj.size = row[4]
            obj.quantity_boxes	= row[5]
          
            w.append(obj)
        #print(e)
        
        return render(request, 'editstock.html', {'ware': w})
    else:
         return render(request, 'emplogin.html')
        
def editsheet(request,d_id):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        d_id=request.POST['d_id']
        date = request.POST['date']
        brandname = request.POST['brandname']
        ml = request.POST['ml']
        ob = request.POST['Openingbal']
        plant = request.POST['plant']
        total = request.POST['total']
        cb = request.POST['Closingbal']
        sale = request.POST['sale']
        rate = request.POST['rate']
        
        
        
        query.execute("update daily_sheet set brand_name ='"+brandname+"',ML='"+ml+"',Opening_bal='"+ob+"',plant='"+plant+"',ob_total='"+total+"' ,Closing_bal='"+cb+"' ,sales='"+sale+"',MRP='"+rate+"'  where d_id='"+d_id+"'")
        conn.commit()
        return redirect('dailysheetemp')
    elif "username" in request.session:
        
        username=request.session['username']
    
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from daily_sheet where d_id='"+d_id+"'")
        result = mycursor.fetchall()
        s=[] 
        for row in result:
            obj= Dailysheet()
            obj.d_id=row[0]
            obj.date=row[1]
            obj.brand_name=row[2] 
            obj.ML= row[3]
            obj.Opening_bal = row[4]
            obj.plant	= row[5]
            obj.ob_total= row[6]  
            obj.Closing_bal= row[7] 
            obj.sales= row[8] 
            obj.MRP= row[9] 
            obj.amount= row[10] 
            s.append(obj)
        
        return render(request, 'editsheet.html',{'sheet':s})
    else:
        
        return render(request, 'emplogin.html')
             
        
def editbrands(request,id):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        id=request.POST['id']
        brn = request.POST['brandnumber']
        brandname = request.POST['brandname']
        ml = request.POST['ML']
        price = request.POST['issue_price']
        mrp = request.POST['MRP']
        
        
        query.execute("update adminapp_brand set brand_no ='"+brn+"',brand_name='"+brandname+"',ML='"+ml+"',issue_price='"+price+"',MRP='"+mrp+"'  where id='"+id+"'")
        conn.commit()
        return redirect('brandsandpriceadmin')
    elif "username" in request.session:
        
        username=request.session['username']
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from adminapp_brand where id='"+id+"'")
        result = mycursor.fetchall()
        b=[] 
        for row in result:
            obj= Brand()
            obj.id=row[0]
            obj.brand_no=row[1] 
            obj.brand_name= row[2]
            obj.ML = row[3]
            obj.issue_price	= row[4]
            obj.MRP= row[5]  
            b.append(obj)
        #print(e)
        
        return render(request, 'editbrands.html', {'brand': b})
    else:
        return render(request,'adminlogin.html')
        
    
def dailysheetemp(request):
    if 'date' in request.POST:
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        query = conn.cursor()
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select d_id,date,brand_name,ML,Opening_bal,plant,ob_total,Closing_bal,sales,MRP,amount from daily_sheet where date between "'+fromdate+'" and "'+todate+'"  ')
        searchresult=query.fetchall()
        d=[]
        a=[]
        R_sale=0 
        Exp=0
        Bal=0
        cashob=0
        total=0
        recash=0
        cash=0
      
        for row in searchresult: 
            obj= Dailysheet()
            obj.d_id=row[0]
            obj.date=row[1] 
            
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
            R_sale=R_sale+int(obj.amount)    
            d.append(obj)
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select  date, retail_sale,expenditure,balance,cashob,total,recash,handcash from dailysales where date between "'+fromdate+'" and "'+todate+'" ')
        result=query.fetchall()

        for row in result:
            obj=Dailysales()
            obj.date=row[0]
            obj.retail_sale=row[1]
            obj.expenditure=row[2]
            obj.balance=row[3]
            obj.cashob=row[4]
            obj.total=row[5]
            obj.recash=row[6]
            obj.handcash=row[7]
            Exp=Exp+int(obj.expenditure)
            Bal=Bal+int(obj.balance)
            cashob=cashob+int(obj.cashob)
            total=total+int(obj.total)
            recash=recash+int(obj.recash)
            cash=cash+int(obj.handcash)
            a.append(obj)            
        return render(request,'dailysheetemp.html',{"data":d,"Rsale" : R_sale, "EXP":Exp,"bal":Bal,"ob":cashob,"tt":total,"cash":recash,"csh":cash })
    elif 'save' in request.POST:
        fromdate=request.POST['fromdate']
        rsale=request.POST['R_Sale']
        exp=request.POST['Exp']
        bal=request.POST['Bal']
        cashob=request.POST['cashob']
        total=request.POST['total']
        recash=request.POST['recash']
        cash=request.POST['cash']
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        query = conn.cursor()
      
        query.execute("insert into dailysales (date,retail_sale,expenditure,balance,cashob, total, recash,handcash)  values('"+fromdate+"','"+rsale+"','"+exp+"','"+bal+"','"+cashob+"', '"+total+"', '"+recash+"', '"+cash+"') ")
        conn.commit()
        return redirect('dailysheetemp')

    
    elif "username" in request.session:
        
        username=request.session['username']
     
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from daily_sheet")
        result = mycursor.fetchall()
        d=[]
        R_sale=0 
        for row in result:  
            obj= Dailysheet()
            obj.d_id=row[0]
            obj.date=row[1] 
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
           
            R_sale=R_sale+int(obj.amount)
            
            d.append(obj)
        #print(e)
        
        return render(request, 'dailysheetemp.html', {'data':d, 'Rsale' : R_sale })
    else:
         return render(request, 'emplogin.html')
 
def dailysheetadmin(request):

    if request.method=='POST':
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        query = conn.cursor()
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select d_id,date,brand_name,ML,Opening_bal,plant,ob_total,Closing_bal,sales,MRP,amount from daily_sheet where date between "'+fromdate+'" and "'+todate+'"  ')
        searchresult=query.fetchall()
        d=[]
        a=[]
        R_sale=0 
        Exp=0
        Bal=0
        cashob=0
        total=0
        recash=0
        cash=0
      
        for row in searchresult: 
            obj= Dailysheet()
            obj.date=row[1] 
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
            R_sale=R_sale+int(obj.amount)    
            d.append(obj)
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        query.execute('select  date, retail_sale,expenditure,balance,cashob,total,recash,handcash from dailysales where date between "'+fromdate+'" and "'+todate+'" ')
        result=query.fetchall()

        for row in result:
            obj=Dailysales()
            obj.date=row[0]
            obj.retail_sale=row[1]
            obj.expenditure=row[2]
            obj.balance=row[3]
            obj.cashob=row[4]
            obj.total=row[5]
            obj.recash=row[6]
            obj.handcash=row[7]
            Exp=Exp+int(obj.expenditure)
            Bal=Bal+int(obj.balance)
            cashob=cashob+int(obj.cashob)
            total=total+int(obj.total)
            recash=recash+int(obj.recash)
            cash=cash+int(obj.handcash)

            a.append(obj)
            

        return render(request,'dailysheetadmin.html',{"data":d,"Rsale" : R_sale, "EXP":Exp,"bal":Bal,"ob":cashob,"tt":total,"cash":recash,"csh":cash})
        
    elif  "username" in request.session:
        
        username=request.session['username']
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
            
        mycursor = conn.cursor()
        mycursor.execute("select * from daily_sheet")
        result = mycursor.fetchall()
        d=[] 
        R_sale=0
        for row in result:
            obj= Dailysheet()
            obj.d_id=row[0]
            obj.date=row[1]
            obj.brand_name	 = row[2]
            obj.ML = row[3]
            obj.Opening_bal	= row[4]
            obj.plant= row[5]
            obj.ob_total= row[6]
            obj.Closing_bal	= row[7]
            obj.sales= row[8]
            obj.MRP= row[9]
            obj.amount=str(int(row[8])*int(row[9])) 
           
            R_sale=R_sale+int(obj.amount)  
          
            d.append(obj)
        #print(e)
        
        return render(request, 'dailysheetadmin.html', {'data': d, 'Rsale' : R_sale})

    else:
        return render(request,'adminlogin.html')
    
def deletestock(request , brand_no):
    if request.method=='POST':
        obj.delete()
        return render(request,"stockmanagementemp.html")
    
def deletesheet(request , brand_name):
    if request.method=='POST':
        obj.delete()
        return render(request,"dailysheet.html")

    
def deletebrands(request,id):  
    conn=mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="inventorymanagement"    
    )
    query=conn.cursor()
    query.execute("DELETE FROM adminapp_brand where id='"+id+"'")
    conn.commit()
    return redirect('brandsandpriceadmin')

    return  redirect('brandsandpriceadmin.html')

def addfiles(request):
    if request.method=='POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        date=request.POST['date']
        query = conn.cursor()
        query.execute("insert into challan_details(date,bill,fileurl) values('"+date+"', '"+filename+"','"+uploaded_file_url +"')")
        conn.commit()
        return render(request, 'addfiles.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'addfiles.html')


def password_recovery(request):

    if request.method == 'POST':
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="inventorymanagement"
        )
        mycursor = conn.cursor()
        # retrive post details
        email = request.POST['email']

        mycursor.execute("select password from employee where email='"+email+"'")

        result = mycursor.fetchone()
        pwd=str(result)
        if (result != None):
            # SMTP server configuration
            smtp_server = 'smtp.gmail.com'
            smtp_port = 587
            smtp_username = 'sathwikakatkuri@gmail.com'
# for App Password enable 2-step verification then u can create app password
            smtp_password = 'nnui cata bgti eooe'

# Email content
            subject = 'Password recovery'
            body = 'This is a Password recovery email sent from SST Wines.'+'Your password as per registration is: '+ pwd[2:len(pwd)-3]
            sender_email = 'sathwikakatkuri@gmail.com'
            receiver_email = email

# Create a message
            message = MIMEMultipart()
            message['From'] = sender_email
            message['To'] = receiver_email
            message['Subject'] = subject
            message.attach(MIMEText(body, 'plain'))

    # Connect to SMTP server and send the email
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_username, smtp_password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            
            return render(request, 'password_recovery.html', {'status': 'Password sent to given mail ID'})
        else:
            return render(request, 'password_recovery.html', {'status': 'Wrong Username!'})
    else:
        return render(request, 'password_recovery.html')  

def password(request):

    if request.method == 'POST':
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="inventorymanagement"
        )
        mycursor = conn.cursor()
        # retrive post details
        email = request.POST['email']

        mycursor.execute("select password from admin where email='"+email+"'")

        result = mycursor.fetchone()
        pwd=str(result)
        if (result != None):
            # SMTP server configuration
            smtp_server = 'smtp.gmail.com'
            smtp_port = 587
            smtp_username = 'sathwikakatkuri@gmail.com'
# for App Password enable 2-step verification then u can create app password
            smtp_password = 'nnui cata bgti eooe'

# Email content
            subject = 'Password recovery'
            body = 'This is a Password recovery email sent from SST Wines.'+'Your password as per registration is: '+ pwd[2:len(pwd)-3]
            sender_email = 'sathwikakatkuri@gmail.com'
            receiver_email = email

# Create a message
            message = MIMEMultipart()
            message['From'] = sender_email
            message['To'] = receiver_email
            message['Subject'] = subject
            message.attach(MIMEText(body, 'plain'))

    # Connect to SMTP server and send the email
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_username, smtp_password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            
            return render(request, 'password.html', {'status': 'Password sent to given mail ID'})
        else:
            return render(request, 'password.html', {'status': 'Wrong Username!'})
    else:
        return render(request, 'password.html')        
             
"""def addbrands(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        # load excel with its path
        BASE_DIR = Path(__file__).resolve().parent.parent
        print(BASE_DIR)
        filepath=str(BASE_DIR)+"\\media\\"+filename
        print(BASE_DIR)
        wrkbk = openpyxl.load_workbook(filepath)
        sh = wrkbk.active
        con=mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="inventorymanagement"
        )
         # iterate through excel and display data
        for i in range(2, sh.max_row+1):
            cell_obj = sh.cell(row=i, column=1)
            bno=cell_obj.value
            cell_obj = sh.cell(row=i, column=2)
            bname=(cell_obj.value)
            cell_obj = sh.cell(row=i, column=3)
            ml=cell_obj.value
            cell_obj = sh.cell(row=i, column=4)
            issue_price=cell_obj.value
            cell_obj = sh.cell(row=i, column=5)
            mrp=cell_obj.value
        newcur=con.cursor()
        newcur.execute("insert into brands(brand_no,brand_name,ML,issue_price,MRP)values('"+str(bno)+"','"+bname+"','"+str(ml)+"','"+str(issue_price)+"','"+str(mrp)+"' ")
        con.commit()
        return render(request, 'addbrands.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'addbrands.html')"""


def removeemp(request, emp_id):
    conn=mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="inventorymanagement"    
    )
    query=conn.cursor()
    query.execute("DELETE FROM employee where emp_id='"+emp_id+"'")
    conn.commit()
    return redirect('others')

def removeacc(request, acc_id):
    conn=mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="inventorymanagement"    
    )
    query=conn.cursor()
    query.execute("DELETE FROM accountant where acc_id='"+acc_id+"'")
    conn.commit()
    return redirect('others')

    
def editemp(request, emp_id):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        username = request.POST['username']
        email = request.POST['email']
        dob = request.POST['dob']
        mobile = request.POST['mobile']
        
        
        query.execute("update employee set username ='"+username+"',email='"+email+"',dob='"+dob+"',mobile='"+mobile+"'  where emp_id='"+emp_id+"'")
        conn.commit()
        return redirect('others')
    elif "username" in request.session:
        
        username=request.session['username']
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from employee where emp_id='"+emp_id+"'")
        result = mycursor.fetchall()
        e=[] 
        for row in result:
            obj= Employee()
            obj.emp_id=row[0]
            obj.username=row[1] 
            obj.email= row[3]
            obj.dob = row[4]
            obj.mobile	= row[5]  
            e.append(obj)
        #print(e)
        
        return render(request, 'editemp.html', {'employee': e})
    else:
        return render(request,'adminlogin.html')

 
def editacc(request, acc_id):
    if request.method=='POST':
        conn = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'inventorymanagement'
        )
        query = conn.cursor()
        username = request.POST['username']
        email = request.POST['email']
        dob = request.POST['dob']
        mobile = request.POST['mobile']
        
        
        query.execute("update accountant set username ='"+username+"',email='"+email+"',dob='"+dob+"',mobile='"+mobile+"'  where acc_id='"+acc_id+"'")
        conn.commit()
        return redirect('others')
    elif "username" in request.session:
        
        username=request.session['username']
            
        conn= mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='inventorymanagement'
        )
        mycursor = conn.cursor()
        mycursor.execute("select * from accountant where acc_id='"+acc_id+"'")
        result = mycursor.fetchall()
        a=[] 
        for row in result:
            obj= Accountant()
            obj.acc_id=row[0]
            obj.username=row[1] 
            obj.email= row[3]
            obj.dob = row[4]
            obj.mobile	= row[5]  
            a.append(obj)
        #print(e)
        
        return render(request, 'editacc.html', {'accountant': a})
    else:
        return render(request,'adminlogin.html')

    
    

        

    


  
    














    


